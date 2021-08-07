import numpy as np
import random
from openpyxl import load_workbook

capacity = 1.8
charge_limit = 0.6


def adjust(value, prev):
    if (value-prev) > charge_limit :
        value = prev+charge_limit
    elif (value-prev) < -charge_limit :
        value = prev-charge_limit

    if value < 0:
        value = 0.0
    elif value > capacity:
       value = capacity

    return value


def make_chromosome():
    # 24 hours + fitness
    chromosome = [0] * 25
   
    # s = np.random.normal(0.3, 0.15, 24) * capacity
    s = np.random.uniform(0.7, 1.3, 24)
    for i in range(len(chromosome) - 1):
        chromosome[i] = adjust(s[i], s[(i+23) % 24])

    return chromosome


def make_population(pop_n):
    population = []

    for n in range(pop_n):
        population.append(make_chromosome())

    return population


def crossover(chromo_1, chromo_2, point_1, point_2):
    chromosome = chromo_1[0:point_1] + chromo_2[point_1:point_2] + chromo_1[point_2:]

    return chromosome


def mutation(chromo, prob):
    length = len(chromo)
    indexes = random.sample(range(length), int(length * prob))

    for index in indexes:
        chromo[index] = chromo[index] + random.uniform(-0.2, 0.2)
        
    for index in range(len(chromo)-1):
        chromo[index] = adjust(chromo[index], chromo[(index+23) % 24])
        
    return chromo


def selection(population):
    sum_of_fitness = 0

    for fitness in population:
        sum_of_fitness += fitness[-1]

    selection_list = []
    for count in range(2):
        fitness_sum = 0
        point = random.uniform(0, sum_of_fitness)
        for chromo in enumerate(population):
            fitness_sum += chromo[1][-1]
            if point < fitness_sum:
                selection_list.append(population[chromo[0]])
                break

    return selection_list


def replacement(population, num_elitism):
    next_population = []

    len_of_pop = len(population)
    population = sorted(population, key=lambda x: x[-1])

    for i in range(num_elitism):
        next_population.append(population[i])

    for j in range(len_of_pop - num_elitism):
        # selected_chromo = selection(population[num_elitism:])
        point_1 = random.randint(0, len_of_pop-1)
        point_2 = random.randint(0, len_of_pop-1)
        new_chromo = crossover(population[point_1], population[point_2], 8, 20)
        new_chromo = mutation(new_chromo, 0.3)
        next_population.append(new_chromo)

    return next_population


def load_dataset(case_name, size):
    load_wb = load_workbook('./ess_data.xlsx', data_only=True)
    load_wb = load_wb[case_name]

    p = [0] * 24
    l = [0] * 24
    g = [0] * 24

    p_asterisk = load_wb['A28'].value

    for i in range(size):
        p[i] = load_wb.cell(i + 2, 1).value
        l[i] = load_wb.cell(i + 2, 2).value
        g[i] = load_wb.cell(i + 2, 3).value

    return p, l, g, p_asterisk


def set_fitness(chromo, p, l, g, p_asterisk):
    value_list = []

    for i in range(len(chromo)-1):
        prev = chromo[(i + 23) % 24]

        value = chromo[i] - prev
        value = value + l[i] - g[i]
        if value < 0 :
            value = 0.0
        value_list.append(value)

    max_value = max(value_list)

    sum_of_value = 0
    cost = 0
    for step in range(len(value_list)):
        sum_of_value += value_list[step] * p[step]
        cost += value_list[step] * p[step]

    sum_of_value += max_value * p_asterisk

    chromo[-1] = sum_of_value

    return chromo


population = make_population(100)
p, l, g, p_asterisk = load_dataset('case01', 24)
for step in range(2000):
    for i in range(len(population)):
        population[i] = set_fitness(population[i], p, l, g, p_asterisk)
    population = replacement(population, 20)
    print('generation ' + str(step + 1) + ' finish')
    print(population[0][-1])

print(population[0])
