import random
from openpyxl import load_workbook


def make_chromosome():
    # 24 hours + fitness
    chromosome = [0] * 25

    for i in range(len(chromosome) - 1):
        chromosome[i] = random.uniform(0, 1.8)

    return chromosome


def make_population(pop_n):
    population = []

    for n in range(pop_n):
        population.append(make_chromosome())

    return population


def crossover(chromo_1, chromo_2, point_1, point_2):
    chromo_1[point_1:point_2] = chromo_2[point_1:point_2]

    return chromo_1


def mutation(chromo, prob):
    if random.uniform(0, 1) > prob:
        random_element = random.choice(chromo[:-1])
        chromo[chromo.index(random_element)] = random_element + random.uniform(-0.1, 0.1)

    return chromo


def selection(population):
    sum_of_fitness = 0

    for fitness in population:
        sum_of_fitness += fitness[-1]

    selection_list = []
    for count in range(2):
        fitness_sum = 0
        point = random.uniform(0, sum_of_fitness)
        for chromo in enumerate(population):
            fitness_sum += chromo[1][-1]
            if point < fitness_sum:
                selection_list.append(population[chromo[0]])
                break

    return selection_list


def replacement(population, num_elitism):
    next_population = []

    len_of_pop = len(population)
    population = sorted(population, key=lambda x: x[-1])

    for i in range(num_elitism):
        next_population.append(population[i])

    for j in range(len_of_pop - num_elitism):
        selected_chromo = selection(population[num_elitism:])
        # point_1 = random.randint(5, 12)
        # point_2 = random.randint(13, 20)
        # new_chromo = crossover(selected_chromo[0], selected_chromo[1], point_1, point_2)
        new_chromo = crossover(selected_chromo[0], selected_chromo[1], 12, 20)
        new_chromo = mutation(new_chromo, 0.7)
        next_population.append(new_chromo)

    return next_population


def set_fitness(chromo):
    load_wb = load_workbook('./ess_data.xlsx', data_only=True)
    load_wb = load_wb['case02']

    p_asterisk = load_wb['A28'].value
    value_list = []

    for cell in enumerate(chromo[:-1]):
        p = load_wb.cell(cell[0] + 2, 1).value
        l = load_wb.cell(cell[0] + 2, 2).value
        g = load_wb.cell(cell[0] + 2, 3).value

        if cell[0] == 0:
            value_list.append(cell[1] + l - g)
        else:
            value_list.append(cell[1] - chromo[cell[0] - 1] + l - g)

    max_value = max(value_list)

    sum_of_value = 0
    for step in enumerate(value_list):
        p = load_wb.cell(step[0] + 2, 1).value
        if step[1] > 0:
            value = step[1] * p + max_value * p_asterisk - (chromo[23] * load_wb.cell(2, 1).value)
            sum_of_value += value
        else:
            value = max_value * p_asterisk - (chromo[23] * load_wb.cell(2, 1).value)
            sum_of_value += value

    chromo[-1] = sum_of_value

    return chromo


population = make_population(100)
for step in range(2000):
    for chromo in enumerate(population):
        population[chromo[0]] = set_fitness(chromo[1])
    population = replacement(population, 10)
    print('generation ' + str(step + 1) + ' finish')
    print(population[0][-1])

print(population[0])
